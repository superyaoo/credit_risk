from datetime import datetime, timedelta
import pandas as pd
from sklearn.metrics import roc_auc_score, roc_curve
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from toollib.model_monitor.online_tools_dories import loan_score_doris2
from toollib.asystem_env.sample_util import get_loans
import time

class ModelReport:
    def __init__(self):
        self.exchange_map = {
            'ac':960,
            'am':20,
            'aec':1,
            'ath':33,
            'athl':33,
            'af':16446
        }
    
    def get_cut_bins(self, df: pd.DataFrame, feature: str, bins: int = 10) -> tuple[pd.Series, list]:
        """将特征值切分成等频分箱
        
        Returns:
            tuple: (分箱结果, 分箱边界值列表)
        """
        print(f"数据总量: {len(df)}")
        print(f"特征 {feature} 非空值数量: {df[feature].notna().sum()}")
        print(f"特征 {feature} 唯一值数量: {df[feature].nunique()}")
        print(f"特征 {feature} 唯一值: {sorted(df[feature].unique())}")
        print(f"请求的分箱数量: {bins}")

        bins_result = pd.qcut(df[feature], bins, duplicates='drop')
        
        print(f"实际创建的分箱数量: {len(bins_result.cat.categories)}")
        print(f"分箱类别: {bins_result.cat.categories}")

        
        bin_edges = sorted([interval.left for interval in bins_result.cat.categories] + [bins_result.cat.categories[-1].right])
        return bin_edges

    def get_model_report(self, 
                         sys_name: str,
                         df: pd.DataFrame, 
                        feature: str, 
                        target: str,
                        user_type: str,  
                        cut_bins: int | list = 10, 
                        lamb: float = 0.001) -> pd.DataFrame:
        """生成模型报告
        
        Args:
            cut_bins: 整数表示等频分箱数量，列表表示自定义分箱边界
        """
        feature_bin = f"{feature}_bin"
        df = df.copy()
        
        if isinstance(cut_bins, int):
            bins_result, _ = self.get_cut_bins(df, feature, cut_bins)
            df[feature_bin] = bins_result
        else:
            df[feature] = pd.to_numeric(df[feature])
            df[feature_bin] = pd.cut(df[feature], cut_bins, duplicates='drop', include_lowest=True)
        
        t_t = df[target].count()
        print('总样本数：{}'.format(t_t))
        b_t = df[target].sum()
        g_t = t_t - b_t
        br = b_t / t_t
        
        dti = df.groupby(feature_bin).agg(
            bad=(target, 'sum'),
            total=(target, 'count'),
            brate=(target, 'mean'),
            installment_amount_sum=('installment_amount', 'sum'),
            installment_amount_mean=('installment_amount', 'mean'),
        )
        dti = dti.sort_values(feature_bin, ascending=False)
        
        bad_amount = df[df[target] == 1].groupby(feature_bin)['installment_amount'].sum()
        good_amount = df[df[target] == 0].groupby(feature_bin)['installment_amount'].sum()
        dti['bad_amount_sum'] = bad_amount / self.exchange_map.get(sys_name, 1)
        dti['good_amount_sum'] = good_amount / self.exchange_map.get(sys_name, 1)
        dti['bad_amount_rate'] = bad_amount / (bad_amount + good_amount)
        dti['installment_amount_sum'] = dti['installment_amount_sum'] / self.exchange_map.get(sys_name, 1)
        dti['installment_amount_mean'] = dti['installment_amount_mean'] / self.exchange_map.get(sys_name, 1)

        dti['total_rate'] = dti['total'] / t_t
        dti['good'] = dti['total'] - dti['bad']
        dti['lift'] = dti['brate'] / br
        dti["woe"] = np.log(
            ((dti["good"] / g_t) + lamb) / ((dti["bad"] / b_t) + lamb)
        )
        dti['good_cum'] = dti["good"].cumsum()
        dti['bad_cum'] = dti["bad"].cumsum()
        dti['brate_cum'] = dti['bad_cum'] / (dti['good_cum'] + dti['bad_cum'])

        dti["ks"] = np.abs((dti["bad_cum"] / b_t) - (dti["good_cum"] / g_t))
        dti["iv"] = (dti["good"] / g_t - dti["bad"] / b_t) * dti["woe"]
        dti['iv'] = dti['iv'].sum()

        dti = dti.reset_index()
        dti.columns.name = None
        
        dti[feature_bin] = dti[feature_bin].astype(str)
        
        n_bins = len(dti)
        dti['bin_code'] = list(range(n_bins))

        def _cum_calc_auc(n):
            mask = df[feature_bin].cat.codes <= n
            df_new = df[mask]
            
            if len(df_new) < 2 or len(df_new[target].unique()) < 2:
                return np.nan
            
            try:
                auc = roc_auc_score(df_new[target], df_new[feature])
                return max(auc, 1 - auc)
            except ValueError:
                return np.nan

        dti['auc_cum'] = dti['bin_code'].map(_cum_calc_auc)
        
        dti['auc_cum'] = dti['auc_cum'].fillna(method='ffill')
        
        dti.rename({feature_bin: 'bin'}, axis=1, inplace=True)
        dti.insert(0, 'score_field', [feature] * dti.shape[0])
        dti.insert(1, 'user_type', [user_type] * dti.shape[0])
        dti.insert(2, 'label', [target] * dti.shape[0])
        model_report = dti[['score_field', 'user_type', 'label', 'bin', 'total', 'total_rate', 'good', 'bad',
            'installment_amount_sum', 'installment_amount_mean', 'good_amount_sum', 'bad_amount_sum', 
            'brate', 'lift', 'brate_cum', 'bad_amount_rate', "woe", 'iv', 'ks', 'auc_cum']]
        summary_row = {
        'score_field': '小计',
        'user_type': '',
        'label': '',
        'bin': '',
        'total': model_report['total'].sum(),
        'total_rate': model_report['total_rate'].sum(),
        'good': model_report['good'].sum(),
        'bad': model_report['bad'].sum(),
        'installment_amount_sum': model_report['installment_amount_sum'].sum(),
        'installment_amount_mean': model_report['installment_amount_sum'].sum() / model_report['total'].sum(),
        'good_amount_sum': model_report['good_amount_sum'].sum(),
        'bad_amount_sum': model_report['bad_amount_sum'].sum(),
        'brate': model_report['bad'].sum() / model_report['total'].sum(),
        'lift': '',
        'brate_cum': model_report['bad'].sum() / model_report['total'].sum(),
        'bad_amount_rate': model_report['bad_amount_sum'].sum() / model_report['installment_amount_sum'].sum(),
        'woe': '',
        'iv': model_report['iv'].max(),
        'ks': model_report['ks'].max(),
        'auc_cum': model_report['auc_cum'].iloc[-1]
        }
        
        model_report_with_summary = pd.concat([model_report, pd.DataFrame([summary_row])], ignore_index=True)
        model_report_with_summary.fillna(0, inplace=True)
        print('==================')
        print(model_report_with_summary)
        print('==================')
        return model_report_with_summary
    
    def calculate_score_psi(self,
                        score1: pd.Series, 
                        score2: pd.Series,
                        cut_bins: list,
                        ) -> tuple[float, pd.DataFrame]:
        dist1 = pd.cut(score1, bins=cut_bins, duplicates='drop')
        dist2 = pd.cut(score2, bins=cut_bins, duplicates='drop')
        
        report = pd.DataFrame({
            'bin': dist1.cat.categories.astype(str),
            'benchmark_count': dist1.value_counts(sort=False),
            'compare_count': dist2.value_counts(sort=False)
        }).fillna(0)
        
        report['benchmark_pct'] = report['benchmark_count'] / report['benchmark_count'].sum()
        report['compare_pct'] = report['compare_count'] / report['compare_count'].sum()
        
        min_pct = 0.0001
        report['benchmark_pct'] = report['benchmark_pct'].replace(0, min_pct)
        report['compare_pct'] = report['compare_pct'].replace(0, min_pct)
        
        report['psi_contribution'] = (report['compare_pct'] - report['benchmark_pct']) * \
                                    np.log(report['compare_pct'] / report['benchmark_pct'])
        
        total_psi = report['psi_contribution'].sum()
        
        total_row = pd.DataFrame({
            'bin': ['Total'],
            'benchmark_count': [report['benchmark_count'].sum()],
            'compare_count': [report['compare_count'].sum()],
            'benchmark_pct': [1],
            'compare_pct': [1],
            'psi_contribution': [total_psi]
        })
        
        report = pd.concat([report, total_row], ignore_index=True).fillna(0)
    
        return total_psi, report

    @classmethod
    def get_concat_report(cls, 
                         sys_name: str,
                         data: pd.DataFrame, 
                         sample_periods: list[tuple[str, str]],
                         user_type: str,
                         cut_bins: list[float],
                         score_field: str,
                         label: str) -> pd.DataFrame:
        model_reports = []
        model = cls()
        datas = []
        for i, period in enumerate(sample_periods):
            data_i = data[
                (data['repayment_date'] >= period[0]) & 
                (data['repayment_date'] < period[1])
            ]
            datas.append(data_i)
            
        benchmark_data = datas[-1]
        
        for i, data_i in enumerate(datas):
            model_report = model.get_model_report(
                sys_name,
                data_i,
                score_field,
                label,
                user_type,
                cut_bins=cut_bins
            )
            
            if i == len(datas) - 1:
                model_report['psi'] = 0
            else:
                psi, psi_report = model.calculate_score_psi(
                    benchmark_data[score_field],
                    data_i[score_field],
                    cut_bins
                )
                model_report['psi'] = psi_report['psi_contribution']
            if i == 0:
                model_reports.append(model_report[model_report.columns.tolist()[:4]])
            model_reports.append(model_report[model_report.columns.tolist()[4:]])
        return pd.concat(model_reports, axis=1)

def get_online_data(sys_name,
                    user,
                    passwd,
                    module_name, 
                     score_field, 
                     new_old_user_status, 
                     app_type,
                     start_date,
                     end_date,
                     dblink
                     ):

    loan_score_df = loan_score_doris2(sys_name=sys_name,
                            user=user, 
                            passwd=passwd,
                            module_name=module_name,
                            score_field=score_field,
                            new_old_user_status=new_old_user_status,
                            start_date=start_date,
                            end_date=end_date,
                            dblink=dblink)
    loan_df = get_loans(sys_name=sys_name,
                        user=user,
                        passwd=passwd,
                        dblink=dblink,
                        start_date=start_date,
                        end_date=end_date,
                        new_old_user_status=new_old_user_status)
    loan_df = loan_df[loan_df['device_type']==app_type]
    loan_df['repayment_date'] = loan_df['repayment_date'].astype(str)
    loan_score_df['repayment_date'] = loan_score_df['repayment_date'].astype(str)
    same_cols = set(loan_df.columns.tolist()) & set(loan_score_df.columns.tolist())
    data = pd.merge(loan_df, loan_score_df[list(loan_score_df.columns.difference(same_cols)) + ['app_order_id', 'repayment_date']], on=['app_order_id', 'repayment_date'])
    data = data[data['extension_count'] == 0]
    return data


def calc_sample_period(performance_days: int, sample_period: int) -> int:
    start_date =(datetime.now() - timedelta(days=sample_period + performance_days)).strftime('%Y-%m-%d')
    end_date = (datetime.now() - timedelta(days=performance_days)).strftime('%Y-%m-%d')
    return start_date, end_date

def get_date_ranges(ref_date: str | datetime, n: int, interval_days: int) -> list[tuple[str, str]]:
    if isinstance(ref_date, str):
        ref_date = datetime.strptime(ref_date, '%Y-%m-%d')
    
    ranges = []
    
    for i in range(n):
        end_date = ref_date - timedelta(days=i * interval_days)
        period_start = end_date - timedelta(days=interval_days)
        ranges.append((
            period_start.strftime('%Y-%m-%d'),
            end_date.strftime('%Y-%m-%d')
        ))
    return ranges

def write_df_to_template(df, template_path, output_path, sheet_name='week', start_cell='B4'):
    wb = load_workbook(template_path)
    ws = wb[sheet_name]
    start_column_letter, start_row = coordinate_from_string(start_cell)
    start_column = column_index_from_string(start_column_letter)
    
    df_rows, df_cols = df.shape
    
    original_formats = {}
    original_row_heights = {}
    for row in range(start_row, start_row + df_rows):
        original_row_heights[row] = ws.row_dimensions[row].height
        for col in range(start_column, start_column + df_cols):
            cell = ws.cell(row=row, column=col)
            original_formats[(row, col)] = cell._style
    
    for i, row in enumerate(df.values):
        current_row = start_row + i
        if current_row in original_row_heights and original_row_heights[current_row] is not None:
            ws.row_dimensions[current_row].height = original_row_heights[current_row]
            
        for j, value in enumerate(row):
            cell = ws.cell(row=current_row, column=start_column+j)
            original_format = original_formats.get((current_row, start_column+j))
            cell.value = value
            if original_format:
                cell._style = original_format
    
    wb.save(output_path)
    return ws