import logging
from pathlib import Path

import pandas as pd
import numpy as np
import math
import json
from toollib.model_monitor.online_tools_dories import loan_score_doris2
from toollib.asystem_env.sample_util import get_loans

from tqdm import tqdm
from openpyxl import load_workbook
from toollib.unversal import numerical_univerate
from datetime import datetime, timedelta
import warnings

warnings.filterwarnings("ignore")
import lightgbm as lgb
import xgboost as xgb
from sklearn.metrics import (
    roc_auc_score,
    roc_curve)

logger = logging.getLogger(__name__)


def save_data_to_excel(df, sheet, row_number, col_number):
    for row_num, index in enumerate(df.index, row_number):
        for col_num, col in enumerate(df.keys(), col_number):
            cell = sheet.cell(row=row_num, column=col_num, value=df[col][row_num - row_number])


def calc_auc(y_true, y_score):
    """
    计算评分或者特征的auc
    :param y_true:
    :param y_score:
    :return:
    """
    return roc_auc_score(y_true, y_score)


def calc_ks(y_true, y_score):
    """
    计算评分或者特征的ks结果
    :param y_true:
    :param y_score:
    :return:
    """
    y_true = np.asarray(y_true)
    y_score = np.asarray(y_score)
    fpr, tpr, _ = roc_curve(y_true, y_score)
    ks = np.max(np.abs(fpr - tpr))
    return ks


def cut_bins(x, n_bins=10, method='freq'):
    """
    数值类特征基础的获取分箱bound的方式
    :param x: array_like 待分箱的数据
    :param n_bins: 分箱数量
    :param method: 分箱的方式 默认 'freq' 等频率分箱，'dist' 对应等间距分箱 ,'chi2' 卡方分箱 ,'bestks' 对应best ks分箱
    :return: list
    """
    data = np.array(x, copy=True, dtype=np.float64)
    data = data[~np.isnan(data)]
    if method == 'freq':
        sorted_data = np.sort(data)
        # 计算每个箱子的数据点索引
        indices = np.linspace(0, len(sorted_data) - 1, n_bins + 1, dtype=int)
        bin_edges = sorted_data[indices]
        bin_edges = np.unique(bin_edges)
        if len(bin_edges) < n_bins + 1:
            bin_edges = np.insert(bin_edges, 0, -np.inf)
        else:
            bin_edges[0] = -np.inf
        bin_edges[-1] = np.inf
        return bin_edges
    elif method == 'dist':
        max_v = np.max(data)
        min_v = np.min(data)
        binlen = (max_v - min_v) / n_bins
        bin_edges = [min_v + i * binlen for i in range(n_bins + 1)]
        bin_edges = np.unique(bin_edges)  # #np.unique(bin_edges).astype(float)
        return bin_edges
    else:
        raise ValueError('method must be \'freq\' or \'dist\'')


def make_bin(y_score, bound, num_fillna=-999, cate_fillna=""):
    """
    根据传入的数据对结果映射分箱，自动填充缺失值，并约束上限
    :param y_score: 待分箱的特征
    :param bound: 数值型特征为 list:number, 类别特征为 dict(string:int)
    :param num_fillna: 数值型特征默认填充的方式
    :param cate_fillna: 类别特征默认的填充方式
    :return:
    """
    data = np.asarray(y_score)
    if pd.api.types.is_numeric_dtype(y_score):
        bound = np.asarray(bound)
        data[np.isnan(data)] = num_fillna
        bound_max = np.max(bound)
        data[data > bound_max] = np.max(bound)
        indices = np.digitize(data, bound, right=True)
        return bound[indices]
    else:
        data = pd.Series(data).fillna(cate_fillna)
        keys = bound.keys()
        data = data.map(lambda x: x if x in keys else cate_fillna)
        return np.asarray(pd.Series(data).map(bound).fillna(0))


def calc_psi(x_true, x_pred, num_bins=10, method='freq', lamb=0.001):
    """
    计算特征psi
    :param x_true:
    :param x_pred:
    :param num_bins:
    :param method:
    :return:
    """
    if pd.api.types.is_numeric_dtype(x_true):
        bound = cut_bins(x_true, n_bins=num_bins, method=method)
        x_base_bins = np.array(make_bin(x_true, bound, num_fillna=-999))
        x_pred_bins = np.array(make_bin(x_pred, bound, num_fillna=-999))
    else:
        x_base_bins = x_true
        x_pred_bins = x_pred

    n_base = len(x_true)
    n_pred = len(x_pred)
    b_bins = np.unique(np.concatenate((x_base_bins, x_pred_bins)))
    psi_mapping = []
    for bin in b_bins:
        c_base = np.count_nonzero(x_base_bins == bin)
        c_pred = np.count_nonzero(x_pred_bins == bin)
        pct_base = (c_base + 0.0) / n_base
        pct_pred = (c_pred + 0.0) / n_pred
        csi = (pct_base - pct_pred) * math.log((pct_base + lamb) / (pct_pred + lamb))
        psi_mapping.append((bin, c_base, c_pred, csi))

    psi = sum([x[3] for x in psi_mapping])
    return psi


def calc_iv(y_true, y_score, num_bins=10, method='freq', fillna=-999, lamb=0.001):
    """
    计算iv
    :param y_true: 标签列
    :param y_score: 特征列
    :return:
    """
    y_true = np.asarray(y_true).copy()
    y_score = np.asarray(y_score).copy()
    y_score[np.isnan(y_score)] = fillna

    bound = cut_bins(y_score, n_bins=num_bins, method=method)
    y_pred_bins = np.asarray(make_bin(y_score, bound, num_fillna=fillna))

    N_g = (y_true == 0).sum()
    N_b = (y_true == 1).sum()
    data = np.column_stack((y_true, y_pred_bins))

    iv_arr = []
    bins = np.unique(y_pred_bins)
    for bin in bins:
        data_bin = data[data[:, 1] == bin]
        n_g = (data_bin[:, 0] == 0).sum()
        n_b = (data_bin[:, 0] == 1).sum()
        pct_g = (n_g * 1.0 / N_g)
        pct_b = (n_b * 1.0 / N_b)
        iv = (pct_g - pct_b) * math.log((pct_g + lamb) / (pct_b + lamb))
        iv_arr.append(iv)
    iv = sum(iv_arr)
    return iv


def calc_psi(x_true, x_pred, num_bins=10, method='freq', lamb=0.001):
    """
    计算特征psi
    :param x_true:
    :param x_pred:
    :param num_bins:
    :param method:
    :return:
    """
    if pd.api.types.is_numeric_dtype(x_true):
        bound = cut_bins(x_true, n_bins=num_bins, method=method)
        x_base_bins = np.array(make_bin(x_true, bound, num_fillna=-999))
        x_pred_bins = np.array(make_bin(x_pred, bound, num_fillna=-999))
    else:
        x_base_bins = x_true
        x_pred_bins = x_pred

    n_base = len(x_true)
    n_pred = len(x_pred)
    b_bins = np.unique(np.concatenate((x_base_bins, x_pred_bins)))
    psi_mapping = []
    for bin in b_bins:
        c_base = np.count_nonzero(x_base_bins == bin)
        c_pred = np.count_nonzero(x_pred_bins == bin)
        pct_base = (c_base + 0.0) / n_base
        pct_pred = (c_pred + 0.0) / n_pred
        csi = (pct_base - pct_pred) * math.log((pct_base + lamb) / (pct_pred + lamb))
        psi_mapping.append((bin, c_base, c_pred, csi))

    psi = sum([x[3] for x in psi_mapping])
    return psi


def _grouped_info(df, col, target, lamb=0.001):
    t_t = df[target].count()
    b_t = df[target].sum()
    g_t = t_t - b_t
    br = b_t / t_t
    dti = df.groupby(col).agg(
        bad=(target, 'sum'),
        total=(target, 'count'),
        brate=(target, 'mean')
    )
    dti['total_pct'] = dti['total'] / t_t
    dti['good'] = dti['total'] - dti['bad']
    dti['lift'] = dti['brate'] / br
    dti["woe"] = np.log(
        ((dti["good"] / g_t) + lamb) / ((dti["bad"] / b_t) + lamb)
    )
    dti['good_cum'] = dti["good"].cumsum()
    dti['bad_cum'] = dti["bad"].cumsum()
    dti['brate_cum'] = dti['bad_cum'] / (dti['good_cum'] + dti['bad_cum'])

    dti["ks"] = np.abs((dti["bad_cum"] / b_t) - (dti["good_cum"] / g_t))
    dti["iv"] = (dti["good"] / g_t - dti["bad"] / b_t) * dti["woe"]
    return dti[['bad', 'total', 'total_pct', 'brate', 'lift', 'woe', 'ks', 'iv']]


def univerate(train_df, valid_df, feature_name, target, fillna=-999, bins=10):
    """
    分析数值特征的单变量结果
    :param train_df:训练集(推荐为train和test的并集)
    :param valid_df:验证集（推荐为oot）
    :param feature_name:待分析的特征名称
    :param target:为target的目标(0为好人，1为坏人)
    """
    train_df[feature_name] = train_df[feature_name].fillna(fillna)
    valid_df[feature_name] = valid_df[feature_name].fillna(fillna)
    # _, bins = pd.qcut(train_df[feature_name], bins, duplicates='drop', retbins=True)
    bins = cut_bins(train_df[feature_name], bins, method='freq')
    new_col = f"{feature_name}_bins"
    train_df[new_col] = pd.cut(train_df[feature_name], bins)
    valid_df[new_col] = pd.cut(valid_df[feature_name], bins)
    iv_train = calc_iv(train_df[target], train_df[feature_name])
    iv_valid = calc_iv(valid_df[target], valid_df[feature_name])

    ks_train = calc_ks(train_df[target], train_df[feature_name])
    ks_valid = calc_ks(valid_df[target], valid_df[feature_name])

    csi = calc_psi(train_df[feature_name], valid_df[feature_name])

    bin_train = _grouped_info(train_df, new_col, target)
    bin_train.columns = [f"{v}_train" for v in bin_train.columns]

    bin_valid = _grouped_info(valid_df, new_col, target)
    bin_valid.columns = [f"{v}_valid" for v in bin_valid.columns]

    result_t = pd.concat([bin_train, bin_valid], axis=1)
    result_t.index.name = 'bin'
    result_t = result_t.reset_index()
    result_t['feature'] = feature_name
    result_t.insert(0, 'feature', result_t.pop('feature'))
    return iv_train, iv_valid, ks_train, ks_valid, csi, result_t


def batch_feature_univariate(train, test, oot, feature_list, target, n_bins, method):
    feature_summary = []
    feature_bininfos = {}

    for col in tqdm(feature_list):
        if ('woe' in col) or ('base_v3_2' in col) or ('base_v3_1' in col):  # 兼容常见命名异常的其概况
            model_name = "_".join(col.split("_")[0:5])
        else:
            model_name = "_".join(col.split("_")[0:4])
        bins = cut_bins(train[col], n_bins=n_bins, method=method)
        train_fea_info = feature_univariate(train[target], train[col], bins=bins)
        train_fea_info['feature'] = col
        test_fea_info = feature_univariate(test[target], test[col], bins=bins)
        oot_fea_info = feature_univariate(oot[target], oot[col], bins=bins)

        train_fea_info[['total', 'pct', 'brate', 'woe', 'iv']] = train_fea_info[
            ['total', 'pct', 'brate', 'woe', 'iv']].round(4)
        test_fea_info[['total', 'pct', 'brate', 'woe', 'iv']] = test_fea_info[
            ['total', 'pct', 'brate', 'woe', 'iv']].round(4)
        oot_fea_info[['total', 'pct', 'brate', 'woe', 'iv']] = oot_fea_info[
            ['total', 'pct', 'brate', 'woe', 'iv']].round(4)

        display_col = ['bin', 'bad', 'total', 'pct', 'brate', 'woe', 'iv']

        display_df = train_fea_info[['feature'] + display_col].merge(test_fea_info[display_col], on='bin',
                                                                     how='left').merge(oot_fea_info[display_col],
                                                                                       on='bin', how='left')
        display_df['bin'] = display_df['bin'].astype(str)
        display_df.columns = ['feature', 'bin', 'bad_train', 'total_train', 'pct_train', 'brate_train', 'woe_train',
                              'iv_train', 'bad_test',
                              'total_test', 'pct_test', 'brate_test', 'woe_test', 'iv_test', 'bad_oot', 'total_oot',
                              'pct_oot',
                              'brate_oot', 'woe_oot', 'iv_oot']

        iv_train = train_fea_info['iv'].max()
        iv_test = test_fea_info['iv'].max()
        iv_oot = oot_fea_info['iv'].max()
        csi1 = calc_psi(train[col], test[col])
        csi2 = calc_psi(train[col], oot[col])
        feature_info = {'feature': col,
                        'iv_train': round(iv_train, 4), 'iv_test': round(iv_test, 4), 'iv_oot': round(iv_oot, 4),
                        'csi(train-test)': round(csi1, 4), 'csi(train-oot)': round(csi2, 4), 'model_name': model_name}

        feature_summary.append(feature_info)
        feature_bininfos[col] = display_df
    feature_summary_df = pd.DataFrame(feature_summary)
    feature_summary_df['ratio'] = feature_summary_df['iv_oot'] / feature_summary_df['iv_train']
    feature_summary_df['ratio'] = feature_summary_df['ratio'].round(4)
    return feature_summary_df, feature_bininfos


def feature_report(train_df, valid_df, feature_list, target, save=True, feature_importance_dict=None):
    """
    批量跑特征并生成结果报告
    :param train_df:
    :param valid_df:
    :feature_list: 特征字段集合
    :param target: 训练标签
    :param save: 是否需要生成特征报告
    """

    data_statis = []

    def _data_base_info(df, name):
        rs = {'data': name,
              'total': df[target].count(),
              'good': df[target].count() - df[target].sum(),
              'bad': df[target].sum(),
              'brate': df[target].mean(),
              'start': df['loan_time'].min(),
              'end': df['loan_time'].max()
              }
        data_statis.append(rs)

    _data_base_info(train_df, 'train')
    _data_base_info(valid_df, 'valid')

    data_info_df = pd.DataFrame(data_statis)

    feature_summary = []
    feature_bininfos = []
    for feature in tqdm(feature_list, desc='feature_analysis'):
        if not pd.api.types.is_numeric_dtype(train_df[feature]):
            continue
        iv_train, iv_valid, ks_train, ks_valid, csi, result_t = univerate(train_df, valid_df, feature, target,
                                                                          fillna=-999)
        summary = {"feature_name": feature, "iv_train": iv_train, "iv_valid": iv_valid,
                   "ks_train": ks_train, "ks_valid": ks_valid, "csi": csi}
        feature_summary.append(summary)
        feature_bininfos.append(result_t)
    feature_summary = pd.DataFrame(feature_summary)
    feature_summary['ratio'] = feature_summary['iv_valid'] / feature_summary['iv_train']

    if feature_importance_dict is not None:
        feature_summary['imps'] = feature_summary['feature_name']

    if save:
        wb = load_workbook(Path(__file__).parent / 'feature_report_template.xlsx')
        data_info_sheet = wb["data_info"]
        feature_summary_sheet = wb["feature_summary"]
        feature_bininfo_sheet = wb["feauture_bininfo"]

        save_data_to_excel(data_info_df, data_info_sheet, 11, 2)

        save_data_to_excel(feature_summary, feature_summary_sheet, 2, 1)

        start_index = 2
        for feature_info in feature_bininfos:
            feature_info['bin'] = feature_info['bin'].astype(str)
            save_data_to_excel(feature_info, feature_bininfo_sheet, start_index, 1)
            start_index = start_index + len(feature_info) + 1

        time_str = datetime.now().strftime("%Y%m%d%H%M%S")
        file_name = f'feature_info{time_str}.xlsx'
        wb.save(file_name)
        print(f"特征报告保存文件:{file_name}")
    return feature_summary, feature_bininfos


def model_feature_info(model, importance_type='gain', filter_zero=True, retType='df'):
    """
    分析模型特征及权重
    :param model:
    :param importance_type:
    :parm filter_zero:
    :parm retType:
    :return:
    """
    if isinstance(model, (lgb.Booster, lgb.LGBMRegressor, lgb.LGBMClassifier)):  # 兼容lgb 的 原生和sklearn接口
        if isinstance(model, lgb.Booster):
            booster = model
        else:  # 兼容lgb sklearn 接口
            booster = model.booster_
        fea_rs = pd.DataFrame(
            {"var": booster.feature_name(), "imps": booster.feature_importance(importance_type=importance_type)}
        ).sort_values("imps", ascending=False)
    elif isinstance(model, (xgb.sklearn.XGBClassifier, xgb.sklearn.XGBRegressor, xgb.core.Booster)):  # 兼容xgb的原生和sklearn接口
        if isinstance(model, xgb.core.Booster):
            booster = model
        else:  # 兼容sklearn接口
            booster = model.get_booster()
        fea_rs = pd.DataFrame(list(booster.get_score(importance_type='total_gain').items()),
                              columns=["var", "imps"]).sort_values('imps', ascending=False)
    else:
        raise ValueError('输入的模型必须是 lgb.Booster,lgb.LGBMRegressor,lgb.LGBMClassifier,xgb.sklearn.XGBClassifier, xgb.sklearn.XGBRegressor, xgb.core.Booster ')

    if filter_zero:
        fea_rs = fea_rs[fea_rs['imps'] > 0]
    if retType == 'dict':
        return fea_rs.set_index('var')['imps'].to_dict()
    else:
        return fea_rs


def prob2score(prob):
    """
    prob转化为概率分的工具
    :param prob:
    :return:
    """
    return round(550 - 60 / np.log(2) * np.log(prob / (1 - prob)), 0)


def feature_univariate(y, x, n_bins=10, bins=None, alpha=0.1, num_fill=-999.0, cate_fill=''):
    """
    分箱分析特征
    :param y : array_like
    :param x: array_like
    :param n_bins: int
    :param bins: list
    :param alpha: float
    :param num_fill: float
    :param cate_fill: str
    :return:
    """
    y = np.array(y, copy=True, dtype=np.float64)
    if pd.api.types.is_numeric_dtype(x):
        x = np.array(x, copy=True, dtype=float)
        if bins is None:
            bins = cut_bins(x, n_bins, method='freq')
        data = pd.DataFrame({'x': x, 'y': y})
        data.fillna(num_fill, inplace=True)
        data['bin'] = pd.cut(data['x'], bins=bins, right=True, retbins=False)
        corr = data['x'].corr(data['y'])
        auc = roc_auc_score(data['y'], data['x'] * -1.0) if corr < 0 else roc_auc_score(data['y'], data['x'])
    else:
        data = pd.DataFrame({'x': x, 'y': y})
        if bins is not None:
            data['bin'] = data['x'].map(bins).fillna(cate_fill)
        else:
            data['bin'] = data['x'].fillna(cate_fill)
            corr = 0.1
            auc = 0

    ascending = corr > 0

    br = data['y'].mean()
    dti = data.groupby('bin', observed=False).agg(
        bad=('y', 'sum'),
        total=('y', 'count'),
        brate=('y', 'mean')
    ).sort_values('bin', ascending=ascending).assign(
        good=lambda x: x['total'] - x['bad'],
        pct=lambda x: x['total'] / (x['total'].sum())
    ).assign(
        bad_cum=lambda x: x['bad'].cumsum(),
        good_cum=lambda x: x['good'].cumsum(),
        bad_prime=lambda x: x['bad'] + alpha,
        good_prime=lambda x: x['good'] + alpha,
        lift=lambda x: (x['brate'] / br)
    ).assign(
        brate_cum=lambda x: x['bad_cum'] / (x['bad_cum'] + x['good_cum']),
        bad_cum_raito=lambda x: x['bad_cum'] / (x['bad'].sum()),
        good_cum_raito=lambda x: x['good_cum'] / (x['good'].sum()),
        bad_ratio_prime=lambda x: x['bad_prime'] / (x['bad_prime'].sum()),
        good_ratio_prime=lambda x: x['good_prime'] / (x['good_prime'].sum()),
    ).assign(
        woe=lambda x: np.log(x['bad_ratio_prime'] / x['good_ratio_prime']),
    ).assign(
        ks=lambda x: abs(x['bad_cum_raito'] - x['good_cum_raito']),
        iv=lambda x: (x['bad_ratio_prime'] - x['good_ratio_prime']) * x['woe']
    ).assign(
        iv=lambda x: x['iv'].sum(),
        auc=auc
    ).reset_index()
    return dti[['bin', 'bad', 'total', 'pct', 'brate', 'brate_cum', 'lift', 'woe', 'ks', 'iv', 'auc']]


def model_report2(data_sets, prob, target, model_obj=None, feature_importance_dict=None, time_col='loan_time', score_name='score_v0',
                  p2score_fun=prob2score, id_col='app_order_id'):
    """
    prob转化为概率分的工具
    :param data_sets :list[pd.Dataframe] 数据集
    :param prob:str 模型predict出来的概率
    :param target:str 好坏标签
    :param time_col:str 时间字段，兼容字符串和时间两种格式，目前支持格式：'2024-05-11','2024-05-11 11:56:31'
    :param p2score_fun:  概率转分的计算公式
    :param score_name:  模型名称
    :param model_obj :训练好的模型对象
    :param feature_importance_dict:dict 特征中的模型重要性
    :param id_col: 订单的id字段，默认是app_order_id，用于统计月信息并去重
    :return:
    """
    assert len(data_sets) > 0 and len(data_sets) < 4, f"datas只能是1-3个数据集"

    if (model_obj is None) and (feature_importance_dict is None):
        raise ValueError(f"入参 model_obj 或 feature_importance_dict 必须输入至少一个")

    if model_obj is not None:
        feature_importance_dict = model_feature_info(model_obj, importance_type='gain', retType='dict')

    feature_list = [k for k, v in sorted(feature_importance_dict.items(), key=lambda x: x[1], reverse=True)]

    wb = load_workbook(Path(__file__).parent / 'model_report_template_v2.xlsx')
    model_sumary_sheet = wb['model_sumary']
    model_desc_sheet = wb['model_desc']
    score_bininfo_sheet = wb['score_bininfo']
    feature_summary_sheet = wb["feature_summary"]
    feature_bininfo_sheet = wb["feauture_bininfo"]
    n_bins = 10
    method = 'freq'
    if len(data_sets) == 1:
        train, test, oot = data_sets[0], data_sets[0], data_sets[0]
    elif len(data_sets) == 2:
        train, test, oot = data_sets[0], data_sets[1], data_sets[1]
    elif len(data_sets) == 3:
        train, test, oot = data_sets[0], data_sets[1], data_sets[2]

    ## 计算特征报告格式
    feature_summary_df, feature_bininfos = batch_feature_univariate(train, test, oot, feature_list, target, n_bins,
                                                                    method)
    feature_summary_df['imps'] = feature_summary_df['feature'].map(feature_importance_dict)

    feature_summary_df = feature_summary_df[['imps', 'feature', 'iv_train', 'iv_test', 'iv_oot', 'csi(train-test)',
                                             'csi(train-oot)', 'ratio', 'model_name']]

    # 写 feature_summary
    save_data_to_excel(feature_summary_df, feature_summary_sheet, 2, 1)

    feature_bininfos_list = [feature_bininfos[x] for x in feature_summary_df['feature'].to_list()]

    # 写 feauture_bininfo
    start_index = 3
    for feature_info in feature_bininfos_list:
        feature_info['bin'] = feature_info['bin'].astype(str)
        save_data_to_excel(feature_info, feature_bininfo_sheet, start_index, 1)
        start_index = start_index + len(feature_info) + 1

    score_col = 'socre'
    train[score_col] = train[prob].map(p2score_fun)
    test[score_col] = test[prob].map(p2score_fun)
    oot[score_col] = oot[prob].map(p2score_fun)

    # 生成 score_bininfo 的内容
    train_df = pd.concat([train[[score_col, target]],
                          test[[score_col, target]]], axis=0)
    score_bin1 = cut_bins(train_df[score_col], n_bins=10, method='freq')
    left_tbl = feature_univariate(train_df[target], train_df[score_col], bins=score_bin1)
    right_tbl = feature_univariate(oot[target], oot[score_col], bins=score_bin1)
    tbl1 = left_tbl.merge(right_tbl, how='left', on='bin')
    tbl1['bin'] = tbl1['bin'].astype(str)
    save_data_to_excel(tbl1, score_bininfo_sheet, 3, 1)

    score_bin2 = cut_bins(train_df[score_col], n_bins=10, method='dist')
    left_tbl = feature_univariate(train_df[target], train_df[score_col], bins=score_bin2)
    right_tbl = feature_univariate(oot[target], oot[score_col], bins=score_bin2)
    tbl2 = left_tbl.merge(right_tbl, how='left', on='bin')
    tbl2['bin'] = tbl2['bin'].astype(str)
    save_data_to_excel(tbl2, score_bininfo_sheet, 16, 1)

    score_bin3 = cut_bins(oot[score_col], n_bins=10, method='freq')
    left_tbl = feature_univariate(oot[target], oot[score_col], bins=score_bin3)
    right_tbl = feature_univariate(train_df[target], train_df[score_col], bins=score_bin3)
    tbl3 = left_tbl.merge(right_tbl, how='left', on='bin')
    tbl3['bin'] = tbl3['bin'].astype(str)
    save_data_to_excel(tbl3, score_bininfo_sheet, 29, 1)

    #  model_desc 页面
    model_desc_df = pd.DataFrame([{
        'score_name': score_name,
        'train_ks': calc_ks(train[target], train[prob]),
        'test_ks': calc_ks(test[target], test[prob]),
        'oot_ks': calc_ks(oot[target], oot[prob]),
        'auc_train': calc_auc(train[target], train[prob]),
        'auc_test': calc_auc(test[target], test[prob]),
        'auc_oot': calc_auc(oot[target], oot[prob]),
        'psi': calc_psi(train[score_col], oot[score_col]),
    }])
    save_data_to_excel(model_desc_df, model_desc_sheet, 4, 2)

    if time_col in train.columns:
        total_df = pd.concat([train[[target, time_col, id_col]], test[[target, time_col, id_col]], oot[[target, time_col, id_col]]], axis=0)

        # 如果传入时间则添加月份统计
        month_df = total_df.copy()
        month_df = month_df.drop_duplicates(subset=[id_col], keep='first')
        month_df['month'] = pd.to_datetime(month_df[time_col]).dt.strftime('%Y-%m')
        rs_month = month_df.groupby('month').agg(
            bad=(target, 'sum'),
            total=(target, 'count'),
            bad_rate=(target, 'mean'),
        ).assign(
            good=lambda x: x['total'] - x['bad']
        ).reset_index()[['month', 'total', 'bad', 'good', 'bad_rate']]
        save_data_to_excel(rs_month, model_sumary_sheet, 20, 2)

    else:
        total_df = pd.concat([train[[target, id_col]], test[[target, id_col]], oot[[target, id_col]]], axis=0)

    data_statis = []

    def _data_base_info(df, name):
        rs = {'data': name,
              'total': df[target].count(),
              'ratio': df[target].count() / total_df[target].count(),
              'good': df[target].count() - df[target].sum(),
              'bad': df[target].sum(),
              'brate': df[target].mean(),
              }
        if time_col in df.columns:
            rs['start'] = df[time_col].min()
            rs['end'] = df[time_col].max()
        data_statis.append(rs)

    _data_base_info(train, 'train')
    _data_base_info(test, 'test')
    _data_base_info(oot, 'oot')

    _data_base_info(total_df, 'total')
    data_info_df = pd.DataFrame(data_statis)
    save_data_to_excel(data_info_df, model_sumary_sheet, 11, 2)

    model_name_df = pd.DataFrame({'model_name': [score_name]})
    save_data_to_excel(model_name_df, model_sumary_sheet, 5, 2)

    time_col_df = pd.DataFrame({'time_col': [time_col]})
    save_data_to_excel(time_col_df, model_sumary_sheet, 9, 3)

    time_col_df = pd.DataFrame({'time_col': [time_col]})
    save_data_to_excel(time_col_df, model_sumary_sheet, 18, 3)

    time_str = datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f'model_report_{score_name}_{time_str}.xlsx'
    wb.save(file_name)
    logger.info(f"报告保存文件:{file_name}")


def model_report(train_df, valid_df, feature_list, score, target):
    """
    生成模型报告函数
    :param train_df: 训练数据（建议传入trian,test的合集）
    :param valid_df: 验证数据（建议传入oot）
    :param feature_list: 模型的特征列表
    :param score: 模型分的字段
    :param target: 好坏标签，1为坏客户，0为好客户
    :param feature_importance_dict:
    """

    model_sumary = []
    dict_train = {
        'total': train_df[target].count(),
        'good': train_df[target].count() - train_df[target].sum(),
        'bad': train_df[target].sum(),
        'bad_rate': train_df[target].mean()
    }
    dict_valid = {
        'total': valid_df[target].count(),
        'good': valid_df[target].count() - valid_df[target].sum(),
        'bad': valid_df[target].sum(),
        'bad_rate': valid_df[target].mean()
    }
    dict_total = {
        'total': dict_train['total'] + dict_valid['total'],
        'good': dict_train['good'] + dict_valid['good'],
        'bad': dict_train['bad'] + dict_valid['bad'],
        'bad_rate': (dict_train['bad'] + dict_valid['bad']) * 1.0 / (dict_train['total'] + dict_valid['total'])
    }
    model_sumary.append(dict_train)
    model_sumary.append(dict_valid)
    model_sumary.append(dict_total)
    model_sumary = pd.DataFrame(model_sumary)

    ## 生成 model_desc 页面信息
    tmp_arr = []
    auc_train = calc_auc(train_df[target], train_df[score])
    auc_valid = calc_auc(valid_df[target], valid_df[score])
    ks_train = calc_ks(train_df[target], train_df[score])
    ks_valid = calc_ks(valid_df[target], valid_df[score])
    psi = calc_psi(train_df[score], valid_df[score])
    model_info_dict = {'model_name': score, 'ks_train': ks_train, 'ks_valid': ks_valid, "auc_train": auc_train, "auc_valid": auc_valid, "psi": psi}
    tmp_arr.append(model_info_dict)
    model_desc_info = pd.DataFrame(tmp_arr)

    # 获取特征信息
    feature_summary, feature_bininfos = feature_report(train_df, valid_df, feature_list, target)

    _, socre_bins = pd.qcut(train_df[score], 10, duplicates='drop', retbins=True)
    socre_bins[0] = -np.inf
    socre_bins[len(socre_bins) - 1] = np.inf
    score_bins_train = numerical_univerate(train_df, score, target, labels=socre_bins)
    score_bins_valid = numerical_univerate(valid_df, score, target, labels=socre_bins)
    score_bins_train['bin'] = score_bins_train['bin'].astype(str)
    score_bins_valid['bin'] = score_bins_valid['bin'].astype(str)

    _, socre_bins = pd.cut(train_df[score], 10, retbins=True)
    socre_bins[0] = -np.inf
    socre_bins[len(socre_bins) - 1] = np.inf
    score_bins_train_dist = numerical_univerate(train_df, score, target, labels=socre_bins)
    score_bins_valid_dist = numerical_univerate(valid_df, score, target, labels=socre_bins)
    score_bins_train_dist['bin'] = score_bins_train_dist['bin'].astype(str)
    score_bins_valid_dist['bin'] = score_bins_valid_dist['bin'].astype(str)

    _, socre_bins = pd.qcut(valid_df[score], 10, duplicates='drop', retbins=True)
    socre_bins[0] = -np.inf
    socre_bins[len(socre_bins) - 1] = np.inf
    score_bins_train_valid_bin = numerical_univerate(train_df, score, target, labels=socre_bins)
    score_bins_valid_valid_bin = numerical_univerate(valid_df, score, target, labels=socre_bins)
    score_bins_train_valid_bin['bin'] = score_bins_train_valid_bin['bin'].astype(str)
    score_bins_valid_valid_bin['bin'] = score_bins_valid_valid_bin['bin'].astype(str)

    wb = load_workbook(Path(__file__).parent / 'model_report_template.xlsx')
    model_sumary_sheet = wb['model_sumary']
    model_desc_sheet = wb['model_desc']
    model_bininfo_sheet = wb['model_bininfo']
    feature_summary_sheet = wb["feature_summary"]
    feature_bininfo_sheet = wb["feauture_bininfo"]

    save_data_to_excel(model_sumary, model_sumary_sheet, 10, 3)
    save_data_to_excel(model_desc_info, model_desc_sheet, 4, 2)

    train_cols = ['bin', 'bad', 'total', 'total_rate', 'brate', 'brate_cum', 'lift', 'woe', 'ks', 'iv', 'auc_cum']
    valid_cols = ['bad', 'total', 'total_rate', 'brate', 'brate_cum', 'lift', 'woe', 'ks', 'iv', 'auc_cum']
    save_data_to_excel(score_bins_train[train_cols], model_bininfo_sheet, 3, 1)
    save_data_to_excel(score_bins_valid[valid_cols], model_bininfo_sheet, 3, 1 + len(train_cols))

    save_data_to_excel(score_bins_train_dist[train_cols], model_bininfo_sheet, 16, 1)
    save_data_to_excel(score_bins_valid_dist[valid_cols], model_bininfo_sheet, 16, 1 + len(train_cols))

    save_data_to_excel(score_bins_valid_valid_bin[train_cols], model_bininfo_sheet, 29, 1)
    save_data_to_excel(score_bins_train_valid_bin[valid_cols], model_bininfo_sheet, 29, 1 + len(train_cols))

    save_data_to_excel(feature_summary, feature_summary_sheet, 2, 1)

    start_index = 2
    for feature_info in feature_bininfos:
        feature_info['bin'] = feature_info['bin'].astype(str)
        save_data_to_excel(feature_info, feature_bininfo_sheet, start_index, 1)
        start_index = start_index + len(feature_info) + 1

    time_str = datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f'model_report{time_str}.xlsx'
    wb.save(file_name)
    print(f"报告保存文件:{file_name}")


def model_offset_report(sys, user, passwd, module_name, score_field, feature_list, target, start_date_A, end_date_A, start_date_B, end_date_B, new_old_user_status=[0], dblink='inner'):
    target_set = {'def_pd0', 'def_pd0_amt', 'def_pd1', 'def_pd1_amt', 'def_pd3', 'def_pd3_amt', 'def_pd4', 'def_pd4_amt', 'def_pd7', 'def_pd7_amt', 'def_pd14', 'def_pd14_amt'}
    if target not in target_set:
        print(f'error:target只能选择{target_set}中的一个')
        return

    feature_list = feature_list[:30]

    result_A = loan_score_doris2(sys, user, passwd, module_name, score_field, new_old_user_status=new_old_user_status, start_date=start_date_A, end_date=end_date_A, dblink=dblink)
    label_A = get_loans(sys, user, passwd, dblink, start_date=start_date_A, end_date=end_date_A, new_old_user_status=new_old_user_status)
    result_A = result_A[result_A['extension_count'] == 0]
    label_A = label_A[(label_A['extension_count'] == 0) & (label_A[f'agr_{target.split("_")[1]}'] == 1)]
    data_A = pd.merge(result_A, label_A, on=['app_order_id'])
    data_A['result'] = data_A['result'].apply(lambda x: json.loads(x))
    store_feature_set_A = set(data_A['result'].iloc[0].keys())
    for fea in feature_list:
        if fea not in store_feature_set_A:
            print(f'error:特征{fea}不在数据集A中')
            return
    df_expand_A = pd.concat([
        data_A.drop(columns=['result', score_field]),
        pd.json_normalize(data_A['result'])], axis=1)

    result_B = loan_score_doris2(sys, user, passwd, module_name, score_field, new_old_user_status=new_old_user_status, start_date=start_date_B, end_date=end_date_B, dblink=dblink)
    label_B = get_loans(sys, user, passwd, dblink, start_date=start_date_B, end_date=end_date_B, new_old_user_status=new_old_user_status)
    result_B = result_B[result_B['extension_count'] == 0]
    label_B = label_B[(label_B['extension_count'] == 0) & (label_B[f'agr_{target.split("_")[1]}'] == 1)]
    data_B = pd.merge(result_B, label_B, on=['app_order_id'])
    data_B['result'] = data_B['result'].apply(lambda x: json.loads(x))
    store_feature_set_B = set(data_B['result'].iloc[0].keys())
    for fea in feature_list:
        if fea not in store_feature_set_B:
            print(f'error:特征{fea}不在数据集B中')
            return
    df_expand_B = pd.concat([
        data_B.drop(columns=['result', score_field]),
        pd.json_normalize(data_B['result'])], axis=1)

    monitor_list = [score_field] + feature_list
    feature_report(df_expand_A, df_expand_B, monitor_list, target)
    return